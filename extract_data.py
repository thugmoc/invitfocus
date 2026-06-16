import openpyxl
import pandas as pd
import logging

logger = logging.getLogger(__name__)


def extract_financial_data(excel_file: str) -> dict:
    """Extract financial data from client Excel file."""
    try:
        wb = openpyxl.load_workbook(excel_file, data_only=True)
    except Exception as e:
        raise ValueError(f"Cannot open Excel file '{excel_file}': {e}")

    sheet_names = [s.lower() for s in wb.sheetnames]
    logger.info(f"Sheets found: {wb.sheetnames}")

    def get_sheet(candidates):
        for name in candidates:
            for actual in wb.sheetnames:
                if actual.lower() == name.lower():
                    return wb[actual]
        return None

    data = {}

    bp_sheet = get_sheet(["Bilan", "Balance Sheet", "BP", "Bilan Prévisionnel"])
    if bp_sheet:
        assets = bp_sheet.cell(row=10, column=2).value or 0
        liabilities = bp_sheet.cell(row=20, column=2).value or 0
        data["assets"] = float(assets)
        data["liabilities"] = float(liabilities)
        data["equity"] = data["assets"] - data["liabilities"]
    else:
        logger.warning("No balance sheet tab found — using zeros")
        data.update({"assets": 0, "liabilities": 0, "equity": 0})

    pl_sheet = get_sheet(["Résultats", "P&L", "Compte de résultat", "CR"])
    if pl_sheet:
        revenue = pl_sheet.cell(row=5, column=2).value or 0
        expenses = pl_sheet.cell(row=15, column=2).value or 0
        data["revenue"] = float(revenue)
        data["expenses"] = float(expenses)
        data["profit"] = data["revenue"] - data["expenses"]
        data["margin"] = (data["profit"] / data["revenue"] * 100) if data["revenue"] > 0 else 0
    else:
        logger.warning("No P&L tab found — using zeros")
        data.update({"revenue": 0, "expenses": 0, "profit": 0, "margin": 0})

    treso_sheet = get_sheet(["Trésorerie", "Treasury", "Cash Flow"])
    if treso_sheet:
        data["treasury"] = float(treso_sheet.cell(row=5, column=2).value or 0)
    else:
        data["treasury"] = 0

    logger.info(f"Extracted data: CA={data['revenue']}, Margin={data['margin']:.1f}%")
    return data


def extract_from_dataframe(df: pd.DataFrame) -> dict:
    """Fallback: extract from a flat DataFrame (CSV or single-sheet Excel)."""
    mapping = {
        "revenue": ["CA", "Revenue", "Chiffre d'affaires", "ca_total"],
        "expenses": ["Charges", "Expenses", "Total charges"],
        "profit": ["Résultat", "Profit", "Net profit"],
        "assets": ["Actif total", "Total assets", "Assets"],
        "liabilities": ["Passif total", "Total liabilities", "Liabilities"],
    }
    data = {}
    for key, candidates in mapping.items():
        for col in candidates:
            if col in df.columns:
                data[key] = float(df[col].iloc[0])
                break
        if key not in data:
            data[key] = 0

    data.setdefault("equity", data["assets"] - data["liabilities"])
    data.setdefault("margin", (data["profit"] / data["revenue"] * 100) if data["revenue"] > 0 else 0)
    data.setdefault("treasury", 0)
    return data
